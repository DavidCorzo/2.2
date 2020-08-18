from openpyxl import load_workbook

def get_rows(filename='prueba.xlsx'):
    wb = load_workbook(filename, read_only=True)

    ws = wb['Sheet1']

    boundaries = [
        'B3', # input("Enter the upper-left coordinate: "),
        'E28'  # input("Enter the lowerright coordinate: ")
    ]

    data_rows = []
    for row in ws[boundaries[0]:boundaries[1]]:
        data_cols = []
        for cell in row:
            data_cols.append(cell.value)
        data_rows.append(data_cols)    
    wb.close()
    return data_rows
